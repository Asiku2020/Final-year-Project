import tkinter 
from tkinter import Tk, Label, PhotoImage, filedialog
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
import openpyxl

# Initialize the available parking slots
available_parking_slots = 60

class MyToplevel(tkinter.Toplevel):
    def print_receipt(self, firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status):
        global available_parking_slots  # Access the global variable
        filename = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
        if filename:
            with open(filename, "w") as file:
                file.write("First Name: " + firstname + "\n")
                file.write("Last Name: " + lastname + "\n")
                file.write("Title: " + title + "\n")
                file.write("Reason: " + reason + "\n")
                file.write("Car Status: " + car_status + "\n")
                file.write("Car Type: " + type_car + "\n")
                file.write("# Registration Number: " + reg_number + "\n")
                file.write("# Time Place: " + time_place + "\n")
                file.write("Registration Status: " + registration_status + "\n")
                file.write("Parking Slot: " + str(available_parking_slots) + "\n")  # Add parking slot info
            messagebox.showinfo("Print", "Receipt saved successfully.")

def login_window():
    login_top = tkinter.Toplevel(padx=20, pady=20)
    login_top.title('Login')

    # Set background image
    try:
        image = Image.open("C:/Users/ASIKU/OneDrive/Desktop/FORM/IMG_20240307_155314_919.jpg",)  # Replace with your image path
        background_color = '# 0000'  # Adjust for desired darkness
        photo = ImageTk.PhotoImage(image)

        image_label = tkinter.Label(login_top, image=photo)
        image_label.image = photo  # Keep a reference to the image
        image_label.place(relwidth=1, relheight=1)  # Stretch image to fill window
    except Exception as e:
        print("Error loading image:", e)

    welcome_label = tkinter.Label(login_top, text="WELCOME TO MUNI UNIVERSITY \nDATA SYSTEM!", font=("Times New Roman", 14))
    welcome_label.pack(pady=10)

    username_label = tkinter.Label(login_top, text='Username:')
    username_label.pack(padx=20, pady=10)

    username_entry = tkinter.Entry(login_top)
    username_entry.pack()

    password_label = tkinter.Label(login_top, text='Password:')
    password_label.pack(padx=20, pady=10)

    password_entry = tkinter.Entry(login_top, show='*')
    password_entry.pack()

    def check_login():
        username = username_entry.get()
        password = password_entry.get()

        # Replace with your actual authentication logic (e.g., database check)
        if username == 'admin' and password == '123':
            login_top.destroy()
            window.deiconify()  # Show the main window
        else:
            messagebox.showerror('Login Error', 'Invalid username or password')

    login_button = tkinter.Button(login_top, text='Login', command=check_login)
    login_button.pack(pady=10)

def enter_data():
    global available_parking_slots  # Access the global variable
    accepted = accept_var.get()

    if accepted == 'Accepted':
        # User info
        firstname = first_name_entry.get()
        lastname = last_name_entry.get()

        if firstname and lastname:
            title = title_combobox.get()
            reason = reason_combobox.get()
            car_status = car_status_combobox.get()
            type_car = type_car_entry.get()

            registration_status = reg_status_var.get()
            reg_number = reg_number_entry.get()
            time_place = time_place_combobox.get()

            # Decrease available parking slots by one
            available_parking_slots -= 1

            print('First Name: ', firstname, 'Last Name: ', lastname)
            print('Title: ', title, 'Reason: ', reason, 'Car Status: ', car_status, 'Car Type: ', type_car)
            print('# Registration Number: ', reg_number, '# Time Place: ', time_place)
            print('Registration Status: ', registration_status)
            print('Available Parking Slots:', available_parking_slots)
            print('........................................................')

            filepath = "C:/Users/ASIKU/OneDrive/Desktop/FORM/data.xlsx"

            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["First Name", "Last Name", "Title", "Reason", "Car Status","Car Type", "# Registration Number", "# Time Place", "Registration Status", "Parking Slot"]
                sheet.append(heading)
                workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.append([firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, available_parking_slots])
            workbook.save(filepath)

            receipt_window = MyToplevel()
            receipt_window.print_receipt(firstname, lastname, title, reason, car_status, type_car ,reg_number, time_place, registration_status)
        else:
            messagebox.showwarning(title='Error', message='firstname and lastname are required')
    else:
        messagebox.showwarning(title='Error', message="You haven't accepted the terms")

window = tkinter.Tk()
window.title('Data Entry Form')

# Set the window icon using a JPEG file
icon_image = Image.open("OIP.jpg",)  # Replace "icon.jpg" with the path to your icon file
icon_image = ImageTk.PhotoImage(icon_image)
window.iconphoto(True, icon_image)

# Display the window
window.mainloop()












####################################3Exit Car from Parking
def exit_car(reg_number):
    global total_parking_slots  # Access the global variable

    # Your code to remove the car from the database or any other necessary operation

    # Increase the available parking slots
    total_parking_slots += 1
    print("Parking slot freed:", total_parking_slots)
    # You can update the GUI to reflect the increased available parking slots if necessary


def exit_window():
    exit_top = tkinter.Toplevel(padx=20, pady=20)
    exit_top.title('Exit Car')

    reg_number_label = tkinter.Label(exit_top, text='Registration Number:')
    reg_number_label.pack(padx=20, pady=10)

    reg_number_entry = tkinter.Entry(exit_top)
    reg_number_entry.pack()

    def remove_car():
        reg_number = reg_number_entry.get()

        # Call the exit_car function with the registration number
        exit_car(reg_number)

        tkinter.messagebox.showinfo('Car Exit', f'Car with registration number {reg_number} has exited the parking lot.')

        exit_top.destroy()

    exit_button = tkinter.Button(exit_top, text='Exit Car', command=remove_car)
    exit_button.pack(pady=10)





################################################################
import tkinter 
from tkinter import Tk, Label, PhotoImage, filedialog
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
import openpyxl

class MyToplevel(tkinter.Toplevel):
    def print_receipt(self, firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, parking_slot_number):
        filename = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
        if filename:
            with open(filename, "w") as file:
                file.write("First Name: " + firstname + "\n")
                file.write("Last Name: " + lastname + "\n")
                file.write("Title: " + title + "\n")
                file.write("Reason: " + reason + "\n")
                file.write("Car Status: " + car_status + "\n")
                file.write("Car Type: " + type_car + "\n")
                file.write("# Registration Number: " + reg_number + "\n")
                file.write("# Time Place: " + time_place + "\n")
                file.write("Registration Status: " + registration_status + "\n")
                file.write("Parking Slot Number: " + str(parking_slot_number) + "\n")
            messagebox.showinfo("Print", "Receipt saved successfully.")

def exit_car(reg_number):
    global total_parking_slots  # Access the global variable

    # Your code to remove the car from the database or any other necessary operation

    # Increase the available parking slots
    total_parking_slots += 1
    print("Parking slot freed:", total_parking_slots)
    # You can update the GUI to reflect the increased available parking slots if necessary

def exit_window():
    exit_top = tkinter.Toplevel(padx=20, pady=20)
    exit_top.title('Exit Car')

    reg_number_label = tkinter.Label(exit_top, text='Registration Number:')
    reg_number_label.pack(padx=20, pady=10)

    reg_number_entry = tkinter.Entry(exit_top)
    reg_number_entry.pack()

    def remove_car():
        reg_number = reg_number_entry.get()

        # Call the exit_car function with the registration number
        exit_car(reg_number)

        tkinter.messagebox.showinfo('Car Exit', f'Car with registration number {reg_number} has exited the parking lot.')

        exit_top.destroy()

    exit_button = tkinter.Button(exit_top, text='Exit Car', command=remove_car)
    exit_button.pack(pady=10)

def enter_data():
    # Your existing code for entering data

    # Update the available parking slots
    if total_parking_slots > 0:
        parking_slot_number = total_parking_slots  # Capture parking slot number before decrementing
        total_parking_slots -= 1  # Decrease available slots by one
        print("Parking slot assigned:", total_parking_slots)
        # Update the database or perform any necessary operations

        receipt_window = MyToplevel()
        receipt_window.print_receipt(firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, parking_slot_number)
    else:
        tkinter.messagebox.showwarning(title='Parking Full', message="Sorry, parking slots are full.")

window = tkinter.Tk()
window.title('Data Entry Form')
window.withdraw()  # Hide the main window initially

# Login window
# Add your login_window function here

# Set the window icon using a JPEG file
# Add your icon image setting code here

# Create a heading label
# Add your heading label code here

# Create frames and widgets for entering data
# Add your frames and widgets code here

# Saving Course info
# Add your course info frames and widgets code here

# Acceptance Form
# Add your acceptance form frames and widgets code here

# Button to enter data
button = tkinter.Button(window, text='Enter data', command=enter_data)
button.grid(row=3, column=0, sticky='news', padx=20, pady=20)

# Button to exit car
exit_button = tkinter.Button(window, text='Exit Car', command=exit_window)
exit_button.grid(row=4, column=0, sticky='news', padx=20, pady=20)

window.mainloop()




############################################################################
def exit_car(reg_number, exit_time):
    global total_parking_slots  # Access the global variable

    # Your code to remove the car from the main database or any other necessary operation

    # Increase the available parking slots
    total_parking_slots += 1
    print("Parking slot freed:", total_parking_slots)
    # You can update the GUI to reflect the increased available parking slots if necessary

    # Save data to history database
    history_filepath = "C:/Users/ASIKU/OneDrive/Desktop/FORM/history_database.xlsx"

    if not os.path.exists(history_filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["First Name", "Last Name", "Title", "Reason", "Car Status", "Car Type", "# Registration Number", "# Time Place", "Registration Status", "Exit Time"]
        sheet.append(heading)
        workbook.save(history_filepath)
    workbook = openpyxl.load_workbook(history_filepath)
    sheet = workbook.active
    # Add the data of the exiting car to the history database
    sheet.append([firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, exit_time])
    workbook.save(history_filepath)

##############################################
def exit_car(firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, exit_time):
    global total_parking_slots  # Access the global variable

    # Your code to remove the car from the main database or any other necessary operation

    # Increase the available parking slots
    total_parking_slots += 1
    print("Parking slot freed:", total_parking_slots)
    # You can update the GUI to reflect the increased available parking slots if necessary

    # Save data to history database
    history_filepath = "C:/Users/ASIKU/OneDrive/Desktop/FORM/history_database.xlsx"

    if not os.path.exists(history_filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["First Name", "Last Name", "Title", "Reason", "Car Status", "Car Type", "# Registration Number", "# Time Place", "Registration Status", "Exit Time"]
        sheet.append(heading)
        workbook.save(history_filepath)
    workbook = openpyxl.load_workbook(history_filepath)
    sheet = workbook.active
    # Add the data of the exiting car to the history database
    sheet.append([firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, exit_time])
    workbook.save(history_filepath)
