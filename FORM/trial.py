import tkinter 
from tkinter import Tk, Label, PhotoImage, filedialog
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
import openpyxl


class MyToplevel(tkinter.Toplevel):
    def print_receipt(self, firstname, lastname, title,reason,car_status,type_car, reg_number, time_place, registration_status,parking_slot_number):
        filename = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
        if filename:
            with open(filename, "w") as file:
                file.write("First Name: " + firstname + "\n")
                file.write("Last Name: " + lastname + "\n")
                file.write("Title: " + title + "\n")
                file.write("Reason: " + reason + "\n")
                file.write("Car Status: " + car_status + "\n")
                file.write("Car Type: " + type_car + "\n")
                file.write("# Registration Number: " + reg_number + "\n")
                file.write("# Time Place: " + time_place + "\n")
                file.write("Registration Status: " + registration_status + "\n")
                file.write("Parking Slot Number:" + str(parking_slot_number) + "\n")
            messagebox.showinfo("Print", "Receipt saved successfully.")


def login_window():
    login_top = tkinter.Toplevel(padx=20, pady=20)
    login_top.title('Login')

    # Set background image
    try:
        image = Image.open("C:/Users/ASIKU/OneDrive/Desktop/FORM/IMG_20240307_155314_919.jpg",)  # Replace with your image path
        background_color = '# 0000'  # Adjust for desired darkness
        photo = ImageTk.PhotoImage(image)

        image_label = tkinter.Label(login_top, image=photo)
        image_label.image = photo  # Keep a reference to the image
        image_label.place(relwidth=1, relheight=1)  # Stretch image to fill window
    except Exception as e:
        print("Error loading image:", e)

    welcome_label = tkinter.Label(login_top, text="WELCOME TO MUNI UNIVERSITY \nDATA SYSTEM!", font=("Times New Roman", 14))
    welcome_label.pack(pady=10)

    username_label = tkinter.Label(login_top, text='Username:')
    username_label.pack(padx=20, pady=10)

    username_entry = tkinter.Entry(login_top)
    username_entry.pack()

    password_label = tkinter.Label(login_top, text='Password:')
    password_label.pack(padx=20, pady=10)

    password_entry = tkinter.Entry(login_top, show='*')
    password_entry.pack()

    def check_login():
        username = username_entry.get()
        password = password_entry.get()

        # Replace with your actual authentication logic (e.g., database check)
        if username == 'admin' and password == '123':
            login_top.destroy()
            window.deiconify()  # Show the main window
        else:
            messagebox.showerror('Login Error', 'Invalid username or password')

    login_button = tkinter.Button(login_top, text='Login', command=check_login)
    login_button.pack(pady=10)


# Define a variable to store the total available parking slots
total_parking_slots = 60

############ EXIT PAGE FUNCTION #########################

def exit_car(firstname, lastname, title, reason, car_status, type_car,reg_number, time_place, registration_status, exit_time):
    global total_parking_slots  # Access the global variable

    # Your code to remove the car from the database or any other necessary operation

    # Increase the available parking slots
    total_parking_slots += 1
    print("Parking slot freed:", total_parking_slots)
    # You can update the GUI to reflect the increased available parking slots if necessary

 # Save data to history database
    history_filepath = "C:/Users/ASIKU/OneDrive/Desktop/FORM/history_database.xlsx"

    if not os.path.exists(history_filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["First Name", "Last Name", "Title", "Reason", "Car Status", "Car Type", "# Registration Number", "# Time Place", "Registration Status", "Exit Time"]
        sheet.append(heading)
        workbook.save(history_filepath)
    workbook = openpyxl.load_workbook(history_filepath)
    sheet = workbook.active
    # Add the data of the exiting car to the history database
    sheet.append([firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, exit_time])
    workbook.save(history_filepath)


def exit_window():
    exit_top = tkinter.Toplevel(padx=20, pady=20)
    exit_top.title('Exit Car')

    reg_number_label = tkinter.Label(exit_top, text='Registration Number:')
    reg_number_label.pack(padx=20, pady=10)

    reg_number_entry_exit = tkinter.Entry(exit_top)  # Entry field for registration number
    reg_number_entry_exit.pack()

    exit_time_label = tkinter.Label(exit_top, text='Exit Time:')
    exit_time_label.pack(padx=20, pady=10)    



    # Creating the exit time selection Combobox
 
    exit_times = ['00:00', '00:15', '00:30', '00:45', '01:00', '01:15', '01:30', '01:45', ...]  # Add more options as needed
    exit_time_combobox = ttk.Combobox(exit_top, values=exit_times)
    exit_time_combobox.pack()

    # Button to confirm car exit
    confirm_exit_button = tkinter.Button(exit_top, text='Confirm Exit', command=lambda: remove_car(reg_number_entry_exit.get(), exit_time_combobox.get()))
    confirm_exit_button.pack(pady=10)

def remove_car(reg_number, exit_time,):
    reg_number = reg_number_entry.get()

    # Get the data of the exiting car
    firstname = first_name_entry.get()
    lastname = last_name_entry.get()
    title = title_combobox.get()
    reason = reason_combobox.get()
    car_status = car_status_combobox.get()
    type_car = type_car_entry.get()
    time_place = time_place_combobox.get()
    registration_status = reg_status_var.get()

    # Call the exit_car function with all the required parameters including exit_time
    exit_car(firstname, lastname, title, reason, car_status, type_car, reg_number, time_place, registration_status, exit_time)

    tkinter.messagebox.showinfo('Car Exit', f'Car with registration number {reg_number} has exited the parking lot.')

    exit_top.destroy()



def enter_data():
    accepted = accept_var.get()
    global total_parking_slots  # Access the global variable

    if accepted == 'Accepted':
        # User info
        firstname = first_name_entry.get()
        lastname = last_name_entry.get()

        if firstname and lastname:
            title = title_combobox.get()
            reason = reason_combobox.get()
            car_status = car_status_combobox.get()
            type_car = type_car_entry.get()

            registration_status = reg_status_var.get()
            reg_number = reg_number_entry.get()
            time_place = time_place_combobox.get()
            # Capture the parking slot number
            parking_slot_number = total_parking_slots
           

            print('First Name: ', firstname, 'Last Name: ', lastname)
            print('Title: ', title, 'Reason: ', reason, 'Car Status: ', car_status, 'Car Type: ', type_car)
            print('# Registration Number: ', reg_number, '# Time Place: ', time_place)
            print('Registration Status: ', registration_status)
            print('Parking Slot Number:', str(parking_slot_number))
            print('........................................................')

            filepath = "C:/Users/ASIKU/OneDrive/Desktop/FORM/carbase.xlsx"

            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["First Name", "Last Name", "Title", "Reason", "Car Status","Car Type", "# Registration Number",
                 "# Time Place", "Registration Status","Parking Slot Number",]
                sheet.append(heading)
                workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.append([firstname, lastname, title, reason, car_status,type_car, reg_number, 
                time_place, registration_status, parking_slot_number])
            workbook.save(filepath)


            # Update the available parking slots

            if total_parking_slots > 0:
                total_parking_slots -= 1  # Decrease available slots by one
                print("Parking slot assigned:", total_parking_slots)
                # Update the database or perform any necessary operations

                receipt_window = MyToplevel()
                receipt_window.print_receipt(firstname, lastname, title, reason, car_status, type_car, reg_number, 
                    time_place, registration_status, parking_slot_number)
            else:
                tkinter.messagebox.showwarning(title='Parking Full', message="Sorry, parking slots are full.")


            
        else:
            tkinter.messagebox.showwarning(title='Error', message='firstname and lastname are required')

    else:
        tkinter.messagebox.showwarning(title='Error', message="You haven't accepted the terms")


window = tkinter.Tk()
window.title('Data Entry Form')
window.withdraw()  # Hide the main window initially

# Login window
login_window()

# Set the window icon using a JPEG file
icon_image = Image.open("OIP.jpg",)  # Replace "icon.jpg" with the path to your icon file
icon_image = ImageTk.PhotoImage(icon_image)
window.iconphoto(True, icon_image)

# Create a heading label
heading_label = Label(window, text="MUNI UNIVERSITY TRACKING AND\n PARKING MANAGEMENT SYSTEM", font=("Times New Roman", 18))
heading_label.grid(row=0, column=0, padx=10, pady=10)

frame = tkinter.Frame(window)
frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

user_info_frame = tkinter.LabelFrame(frame, text='User Information')
user_info_frame.grid(row=0, column=0, padx=25, pady=25)

first_name_label = tkinter.Label(user_info_frame, text='First Name')
first_name_label.grid(row=0, column=0)
last_name_label = tkinter.Label(user_info_frame, text='Last Name')
last_name_label.grid(row=0, column=1)

first_name_entry = tkinter.Entry(user_info_frame)
last_name_entry = tkinter.Entry(user_info_frame)
first_name_entry.grid(row=1, column=0)
last_name_entry.grid(row=1, column=1)

title_label = tkinter.Label(user_info_frame, text='Title')
title_combobox = ttk.Combobox(user_info_frame, values=['', 'Mr. ', 'Dr. ', 'Mrs. ', 'Egneer. '])
title_label.grid(row=0, column=2)
title_combobox.grid(row=1, column=2)

reason_label = tkinter.Label(user_info_frame, text='Reason')
reason_combobox = ttk.Combobox(user_info_frame, values=['', 'Staff', 'Student', 'Management','Vistor'])
reason_label.grid(row=2, column=0)
reason_combobox.grid(row=3, column=0)

car_status_label = tkinter.Label(user_info_frame, text='Car/Moto bike  Status')
car_status_combobox = ttk.Combobox(user_info_frame, values=['Private-car', 'Campus-car','Motor-bike'])
car_status_label.grid(row=2, column=1)
car_status_combobox.grid(row=3, column=1)


type_car_label = tkinter.Label(user_info_frame, text='Type of Car')
type_car_label.grid(row=2, column=2)
type_car_entry = tkinter.Entry(user_info_frame)
type_car_entry.grid(row=3, column=2)

for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Saving Registration infor

reg_frame = tkinter.LabelFrame(frame, text='Registration Status')
reg_frame.grid(row=1, column=0, sticky='news', padx=20, pady=20)

#registered_label = tkinter.Label(courses_frame, text='Registration Status')

reg_status_var = tkinter.StringVar(value='Not Registered')
registered_check = tkinter.Checkbutton(reg_frame, text='Currently Registered', variable=reg_status_var,
                                      onvalue='Registered', offvalue='Not Registered')

#registered_label.grid(row=0, column=0)
registered_check.grid(row=1, column=0)

reg_number_label = tkinter.Label(reg_frame, text='# Registration Number')
reg_number_label.grid(row=0, column=1)
reg_number_entry = tkinter.Entry(reg_frame)
reg_number_entry.grid(row=1, column=1)



################################################################

time_place_label = tkinter.Label(reg_frame, text='# Time Place (24-hour)')
time_place_label.grid(row=0, column=2)


# Creating the 24-hour time selection Combobox


hours = [str(h).zfill(2) for h in range(24)]
minutes = ['00', '15', '30', '45']
time_values = [f"{hour}:{minute}" for hour in hours for minute in minutes]
time_place_combobox = ttk.Combobox(reg_frame, values=time_values)
time_place_combobox.grid(row=1, column=2)




for widget in reg_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Acceptance Form
terms_frame = tkinter.LabelFrame(frame, text='Terms & Conditions')
terms_frame.grid(row=2, column=0, sticky='news', padx=20, pady=20)

# Check Terms
accept_var = tkinter.StringVar(value='Not Accepted')
terms_check = tkinter.Checkbutton(terms_frame, text='I accepted the Terms & Conditions',
                                  variable=accept_var, onvalue='Accepted', offvalue='Not Accepted')
terms_check.grid(row=0, column=0)

# Button
button = tkinter.Button(frame, text='Enter data', command=enter_data)
button.grid(row=3, column=0, sticky='news', padx=20, pady=20)


# Button to exit car
exit_button = tkinter.Button(window, text='Exit Car', command=exit_window)
exit_button.grid(row=4, column=0, sticky='news', padx=20, pady=20)



window.mainloop()
